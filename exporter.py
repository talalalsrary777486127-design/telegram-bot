import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

CATEGORY_STYLES = {
    "telegram":          {"fill": "2CA5E0", "emoji": "📱"},
    "telegram_group":    {"fill": "1A7DAF", "emoji": "👥"},
    "telegram_channel":  {"fill": "3BBCF0", "emoji": "📢"},
    "whatsapp":          {"fill": "25D366", "emoji": "💬"},
    "whatsapp_group":    {"fill": "128C7E", "emoji": "👥"},
    "whatsapp_direct":   {"fill": "34B7F1", "emoji": "💬"},
    "other":             {"fill": "7F8C8D", "emoji": "🔗"},
}

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="center")


def _header_row(ws, headers: list[tuple[str, int]], fill_hex: str):
    fill = PatternFill("solid", fgColor=fill_hex)
    font = Font(bold=True, color="FFFFFF", size=11)
    for col, (header, width) in enumerate(headers, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = BORDER
    ws.row_dimensions[1].height = 22


def _fill_link_rows(ws, links: list[dict], fill_hex: str = None):
    headers = [("#", 6), ("Link", 72), ("Type", 10), ("Source", 10), ("First Seen", 18)]
    _header_row(ws, headers, fill_hex or "2C3E50")

    if not links:
        ws.cell(row=2, column=2, value="No links found").alignment = WRAP
        return

    for i, row in enumerate(links, start=1):
        r = i + 1
        _cell(ws, r, 1, i, CENTER)
        lc = ws.cell(row=r, column=2, value=row["url"])
        lc.alignment = WRAP
        lc.border = BORDER
        if row["url"].startswith("http"):
            lc.hyperlink = row["url"]
            lc.font = Font(color="0563C1", underline="single")
        _cell(ws, r, 3, row.get("subcategory", ""), CENTER)
        _cell(ws, r, 4, row.get("source", ""), CENTER)
        ts = row.get("first_seen_at", "")[:16].replace("T", " ")
        _cell(ws, r, 5, ts, CENTER)
        ws.row_dimensions[r].height = 16


def build_category_excel(links: list[dict], category: str, title: str = None) -> io.BytesIO:
    style_key = category
    style = CATEGORY_STYLES.get(style_key, CATEGORY_STYLES["other"])
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (title or category.replace("_", " ").title())[:31]
    _fill_link_rows(ws, links, style["fill"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = f"{title or category}.xlsx"
    return buf


def build_subcategory_excels(
    tg_groups: list[dict],
    tg_channels: list[dict],
    wa_groups: list[dict],
) -> tuple[io.BytesIO, io.BytesIO, io.BytesIO]:
    """Returns (Telegram_Groups.xlsx, Telegram_Channels.xlsx, WhatsApp_Groups.xlsx)."""
    def make(links, style_key, sheet_title):
        style = CATEGORY_STYLES.get(style_key, CATEGORY_STYLES["other"])
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_title[:31]
        _fill_link_rows(ws, links, style["fill"])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    return (
        make(tg_groups,   "telegram_group",   "Telegram Groups"),
        make(tg_channels, "telegram_channel", "Telegram Channels"),
        make(wa_groups,   "whatsapp_group",   "WhatsApp Groups"),
    )


def build_combined_excel(
    telegram: list[dict],
    whatsapp: list[dict],
    other: list[dict],
) -> io.BytesIO:
    wb = openpyxl.Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    _build_summary_sheet(ws_sum, telegram, whatsapp, other)

    for cat, rows, style_key in [
        ("Telegram", telegram, "telegram"),
        ("WhatsApp", whatsapp, "whatsapp"),
        ("Other",    other,    "other"),
    ]:
        ws = wb.create_sheet(title=cat)
        style = CATEGORY_STYLES[style_key]
        _fill_link_rows(ws, rows, style["fill"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    buf.name = "all_links.xlsx"
    return buf


def _build_summary_sheet(ws, telegram, whatsapp, other):
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 12
    dark_fill = PatternFill("solid", fgColor="2C3E50")
    hfont = Font(bold=True, color="FFFFFF", size=12)
    for col, label in enumerate(["Category", "Count"], start=1):
        c = ws.cell(row=1, column=col, value=label)
        c.font = hfont
        c.fill = dark_fill
        c.alignment = CENTER
        c.border = BORDER
    ws.row_dimensions[1].height = 22

    data = [
        ("📱 Telegram",  len(telegram), "2CA5E0"),
        ("💬 WhatsApp",  len(whatsapp), "25D366"),
        ("🔗 Other",     len(other),    "7F8C8D"),
    ]
    total = sum(d[1] for d in data)
    for i, (label, count, color) in enumerate(data, start=2):
        fill = PatternFill("solid", fgColor=color)
        c1 = ws.cell(row=i, column=1, value=label)
        c1.fill = fill; c1.alignment = WRAP; c1.border = BORDER
        c1.font = Font(color="FFFFFF", bold=True)
        c2 = ws.cell(row=i, column=2, value=count)
        c2.fill = fill; c2.alignment = CENTER; c2.border = BORDER
        c2.font = Font(color="FFFFFF", bold=True)

    total_fill = PatternFill("solid", fgColor="1A252F")
    row = len(data) + 2
    c1 = ws.cell(row=row, column=1, value="Total Unique Links")
    c1.fill = total_fill; c1.alignment = WRAP; c1.border = BORDER
    c1.font = Font(color="FFFFFF", bold=True)
    c2 = ws.cell(row=row, column=2, value=total)
    c2.fill = total_fill; c2.alignment = CENTER; c2.border = BORDER
    c2.font = Font(color="FFFFFF", bold=True)


def build_txt(telegram: list, whatsapp: list, other: list) -> io.BytesIO:
    lines = []
    now = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    lines.append("=" * 60)
    lines.append("LINK EXTRACTION REPORT")
    lines.append(f"Generated: {now}")
    lines.append("=" * 60)
    lines.append("")
    for label, items in [
        (f"TELEGRAM LINKS ({len(telegram)})", telegram),
        (f"WHATSAPP LINKS ({len(whatsapp)})", whatsapp),
        (f"OTHER LINKS ({len(other)})", other),
    ]:
        lines.append(label)
        lines.append("-" * 40)
        lines.extend(items if items else ["(none found)"])
        lines.append("")
    total = len(telegram) + len(whatsapp) + len(other)
    lines.append("=" * 60)
    lines.append(f"TOTAL UNIQUE LINKS: {total}")
    lines.append("=" * 60)
    buf = io.BytesIO("\n".join(lines).encode("utf-8"))
    buf.name = "links.txt"
    return buf


def _cell(ws, row, col, value, alignment=None):
    c = ws.cell(row=row, column=col, value=value)
    c.border = BORDER
    if alignment:
        c.alignment = alignment
    return c
